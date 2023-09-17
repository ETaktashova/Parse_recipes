import openpyxl
import os
import parser_blud
from parser_blud import array
from pathlib import Path
# from openpyxl import styles
# from openpyxl.styles import Alignment


path_to_save = f"{Path.home()}\\Desktop\\recipes.xlsx"
if os.path.exists(path_to_save)==False:
    wb = openpyxl.Workbook()    
    wsheet = wb.create_sheet(input('Как назвать страницу файла с желаемыми рецептами?'))
    del wb['Sheet']
else:
    # тут ошибка
    wb = openpyxl.load_workbook(path_to_save)    
    wsheet = wb.create_sheet(input('Как назвать страницу файла с новым списком рецептов?'))   
wb.save(path_to_save)
wb.close()



def formating_sheet():
    # for sheet in wb:
    # rich_string = openpyxl
    wsheet.append(['Название блюда', 'Время приготовления',
                'Ингредиенты', 'Способ приготовления',
                'Фото', 'Ссылка на рецепт'])
    wsheet.column_dimensions['A'].width = 40
    wsheet.column_dimensions['B'].width = 20
    wsheet.column_dimensions['C'].width = 40
    wsheet.column_dimensions['D'].width = 70
    wsheet.column_dimensions['E'].width = 70
    wsheet.column_dimensions['F'].width = 70
    
    # for i in range(wsheet.max_row):
    #     i=2
    #     wsheet[f'D{i}'].alignment = Alignment(wrap_text=True)
    #     i+=1
      
def write_dishes(param):
    for row in param():  
        wsheet.append(row)
   

if __name__ == '__main__':
    formating_sheet()
    write_dishes(array)
    wb.save(path_to_save)
    wb.close()
